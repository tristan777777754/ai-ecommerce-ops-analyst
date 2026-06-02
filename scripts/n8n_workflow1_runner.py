from __future__ import annotations

import json
import os
import subprocess
import sys
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path("/Users/tristan/AI E-commerce Ops Analyst")
LARK_CLI = Path(os.environ.get("LARK_CLI", "/Users/tristan/.npm-global/bin/lark-cli"))
FOLDER_TOKEN = "RbBdfI3vElN7nCdOGXdjhLySpec"
RAW_WORKBOOK = "outputs/lark_import_staging/olist_raw_tables_for_lark_base.xlsx"
ANALYSIS_WORKBOOK_FULL = "outputs/lark_import_staging/olist_analysis_layer_for_lark_base.xlsx"
ANALYSIS_WORKBOOK = "outputs/lark_import_staging/olist_analysis_layer_compact_for_lark_base.xlsx"

COMPACT_ANALYSIS_SHEETS = [
    "executive_summary",
    "monthly_revenue_trend",
    "rfm_segment_summary",
    "dormant_high_value",
    "seller_performance",
    "seller_risk_watchlist",
    "category_performance",
    "category_opportunity",
    "product_performance_top",
    "delivery_risk_by_state",
    "review_risk_by_category",
]


def mask_sensitive(text: str) -> str:
    replacements = ["access_token", "refresh_token", "app_secret", "authorization"]
    lines: list[str] = []
    for line in text.splitlines():
        lowered = line.lower()
        if any(key in lowered for key in replacements):
            lines.append("[masked sensitive output]")
        else:
            lines.append(line)
    return "\n".join(lines)


def run_command(argv: list[str]) -> dict[str, Any]:
    env = os.environ.copy()
    env["HOME"] = "/Users/tristan"
    env["PATH"] = (
        "/opt/venv/bin:/Users/tristan/.npm-global/bin:/usr/local/bin:"
        f"/opt/anaconda3/bin:{env.get('PATH', '')}"
    )

    completed = subprocess.run(
        argv,
        cwd=ROOT,
        env=env,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=900,
        check=False,
    )
    return {
        "ok": completed.returncode == 0,
        "returncode": completed.returncode,
        "stdout": mask_sensitive(completed.stdout)[-8000:],
        "stderr": mask_sensitive(completed.stderr)[-8000:],
    }


def create_compact_analysis_workbook() -> dict[str, Any]:
    source = ROOT / ANALYSIS_WORKBOOK_FULL
    target = ROOT / ANALYSIS_WORKBOOK
    source_wb = load_workbook(source, read_only=True, data_only=True)
    target_wb = Workbook(write_only=True)
    copied: list[str] = []

    for sheet_name in COMPACT_ANALYSIS_SHEETS:
        if sheet_name not in source_wb.sheetnames:
            continue
        source_ws = source_wb[sheet_name]
        target_ws = target_wb.create_sheet(title=sheet_name)
        for row in source_ws.iter_rows(values_only=True):
            target_ws.append(row)
        copied.append(sheet_name)

    if not copied:
        raise RuntimeError("No analysis sheets were copied into compact workbook")

    target.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(target)
    source_wb.close()
    return {
        "compact_workbook": ANALYSIS_WORKBOOK,
        "compact_size_mb": round(target.stat().st_size / (1024 * 1024), 1),
        "sheets": copied,
    }


def command_for(step: str, run_id: str) -> list[str]:
    if step == "build_raw":
        return [sys.executable, "scripts/build_lark_import_workbook.py"]
    if step == "import_raw":
        return [
            str(LARK_CLI),
            "drive",
            "+import",
            "--type",
            "bitable",
            "--file",
            RAW_WORKBOOK,
            "--name",
            f"AI E-commerce Ops Analyst - Raw Data {run_id}",
            "--folder-token",
            FOLDER_TOKEN,
            "--as",
            "user",
        ]
    if step == "build_analysis":
        return [sys.executable, "scripts/build_lark_analysis_workbook.py"]
    if step == "import_analysis":
        return [
            str(LARK_CLI),
            "drive",
            "+import",
            "--type",
            "bitable",
            "--file",
            ANALYSIS_WORKBOOK,
            "--name",
            f"AI E-commerce Ops Analyst - Analysis Layer {run_id}",
            "--folder-token",
            FOLDER_TOKEN,
            "--as",
            "user",
        ]
    if step == "build_business_recommendations":
        return [sys.executable, "scripts/build_business_recommendations.py"]
    if step == "publish_business_recommendations":
        return [sys.executable, "scripts/publish_business_recommendations_to_lark.py"]
    raise ValueError(f"Unknown step: {step}")


class Handler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        if self.path != "/healthz":
            self.send_json(404, {"ok": False, "error": "not found"})
            return
        self.send_json(200, {"ok": True, "service": "n8n-workflow1-runner"})

    def do_POST(self) -> None:
        if self.path != "/run":
            self.send_json(404, {"ok": False, "error": "not found"})
            return

        try:
            length = int(self.headers.get("content-length", "0"))
            payload = json.loads(self.rfile.read(length) or b"{}")
            step = str(payload.get("step", ""))
            run_id = str(payload.get("run_id", "manual"))
            argv = command_for(step, run_id)
            result = run_command(argv)
            if step == "build_analysis" and result["ok"]:
                result["compact"] = create_compact_analysis_workbook()
            status = 200 if result["ok"] else 500
            self.send_json(status, {"step": step, "run_id": run_id, **result})
        except Exception as exc:
            self.send_json(500, {"ok": False, "error": str(exc)})

    def log_message(self, format: str, *args: Any) -> None:
        return

    def send_json(self, status: int, body: dict[str, Any]) -> None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("content-type", "application/json; charset=utf-8")
        self.send_header("content-length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


if __name__ == "__main__":
    server = ThreadingHTTPServer(("0.0.0.0", 8788), Handler)
    server.serve_forever()
